"""
Consulta Masiva ONPE 2026
Endpoint: POST https://consultaelectoral.onpe.gob.pe/v1/api/consulta/definitiva
Body:      {"dni": "12345678"}
"""

import os
import threading
import uuid

import requests
import urllib3
from flask import Flask, request as flask_request, render_template_string, jsonify, send_file
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = Flask(__name__)

API_URL    = "https://consultaelectoral.onpe.gob.pe/v1/api/consulta/definitiva"
BASE_URL   = "https://consultaelectoral.onpe.gob.pe"
INICIO_URL = "https://consultaelectoral.onpe.gob.pe/inicio"

HEADERS_BASE = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "es-PE,es;q=0.9,en;q=0.8",
}

HEADERS_API = {
    **HEADERS_BASE,
    "Accept":          "application/json, text/plain, */*",
    "Accept-Encoding": "identity",   # sin compresion — recibir JSON plano
    "Content-Type":    "application/json",
    "Origin":          BASE_URL,
    "Referer":         "https://consultaelectoral.onpe.gob.pe/main/local-de-votacion",
}


def _crear_sesion() -> requests.Session:
    """Visita el portal primero para obtener cookies, como lo haría un navegador."""
    s = requests.Session()
    s.verify = False
    # Dejar que requests/urllib3 maneje Accept-Encoding y decompresion automaticamente
    s.headers.update(HEADERS_BASE)
    try:
        r = s.get(INICIO_URL, timeout=15)
        s.get(f"{BASE_URL}/main/local-de-votacion", timeout=10)
    except Exception:
        pass
    return s


_sesion: requests.Session | None = None
_sesion_lock = threading.Lock()


def _get_sesion() -> requests.Session:
    global _sesion
    with _sesion_lock:
        if _sesion is None:
            _sesion = _crear_sesion()
    return _sesion


def _reset_sesion():
    global _sesion
    with _sesion_lock:
        _sesion = None

# Estilos Excel
YES_FILL  = PatternFill("solid", start_color="D4EDDA")
NO_FILL   = PatternFill("solid", start_color="F8D7DA")
ERR_FILL  = PatternFill("solid", start_color="E2E3E5")
YES_FONT  = Font(name="Arial", size=10, bold=True, color="155724")
NO_FONT   = Font(name="Arial", size=10, bold=True, color="721C24")
CELL_FONT = Font(name="Arial", size=10)
HDR_FILL  = PatternFill("solid", start_color="003087")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CENTER    = Alignment(horizontal="center", vertical="center")
LEFT      = Alignment(horizontal="left",   vertical="center")
THIN      = Side(style="thin", color="CCCCCC")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

estado = {"running": False, "total": 0, "done": 0, "errores": 0, "log": []}


# ─── Consulta ONPE ────────────────────────────────────────────────────────────

def consultar_dni(dni: str) -> dict:
    resultado = {
        "ok": False, "es_miembro": False,
        "cargo": "", "region": "", "provincia": "", "distrito": "",
        "mesa": "", "local": "", "direccion": "",
        "pabellon": "", "aula": "", "piso": "",
        "referencia": "", "tipo_voto": "", "error": "",
    }

    sesion = _get_sesion()

    try:
        resp = sesion.post(
            API_URL,
            json={"dni": dni},
            headers=HEADERS_API,
            timeout=20,
        )
    except requests.exceptions.Timeout:
        resultado["error"] = "Tiempo de espera agotado"
        return resultado
    except requests.exceptions.ConnectionError as e:
        _reset_sesion()
        resultado["error"] = f"Error de conexion: {e}"
        return resultado

    try:
        payload = resp.json()
    except ValueError:
        # Reintentar con sesion nueva
        _reset_sesion()
        sesion2 = _get_sesion()
        try:
            resp2   = sesion2.post(API_URL, json={"dni": dni}, headers=HEADERS_API, timeout=20)
            payload = resp2.json()
        except Exception:
            resultado["error"] = (
                f"HTTP {resp.status_code} | "
                f"CT: {resp.headers.get('Content-Type','')} | "
                f"CE: {resp.headers.get('Content-Encoding','')} | "
                f"Hex inicio: {resp.content[:16].hex()}"
            )
            return resultado

    if not payload.get("success", False):
        msg = payload.get("message") or payload.get("titulo") or "Error desconocido"
        resultado["error"] = f"ONPE reporto error: {msg}"
        return resultado

    data = payload.get("data") or {}

    resultado["ok"]         = True
    resultado["es_miembro"] = bool(data.get("miembroMesa", False))

    if resultado["es_miembro"]:
        # Cargo: viene como "ERES SECRETARIO" → quitar prefijo
        cargo_raw = data.get("cargo", "")
        resultado["cargo"] = cargo_raw.replace("ERES ", "").strip().title()

        # Ubigeo: "CALLAO / CALLAO / VENTANILLA"
        ubigeo = data.get("ubigeo", "")
        partes = [p.strip().title() for p in ubigeo.split("/")]
        resultado["region"]    = partes[0] if len(partes) > 0 else ""
        resultado["provincia"] = partes[1] if len(partes) > 1 else ""
        resultado["distrito"]  = partes[2] if len(partes) > 2 else ""

        resultado["mesa"]      = str(data.get("mesaSufragio", ""))
        resultado["local"]     = data.get("localVotacion", "").title()
        resultado["direccion"] = data.get("direccion", "").title()
        resultado["pabellon"]  = str(data.get("pabellon", "") or "")
        resultado["aula"]      = str(data.get("aula", "") or "")
        resultado["piso"]      = str(data.get("piso", "") or "")
        resultado["referencia"]= data.get("referencia", "").title()
        resultado["tipo_voto"] = data.get("tipoVoto", "").title()

    return resultado


# ─── Procesamiento Excel ──────────────────────────────────────────────────────

def procesar_excel(ruta_entrada: str, ruta_salida: str):
    global estado
    estado.update({"running": True, "done": 0, "errores": 0, "log": []})

    def log(msg):
        estado["log"].append(msg)

    wb = load_workbook(ruta_entrada)
    ws = wb.active

    try:
        # Detectar columnas por cabecera
        cols = {
            "nombre": 2, "dni": 3, "estado": 4, "cargo": 5,
            "region": 6, "prov": 7,  "dist": 8,  "mesa": 9,
            "local":  10, "dir": 11,  "ref":  12,  "tipo": 13,
        }
        for cell in ws[1]:
            v = (cell.value or "").strip().lower()
            c = cell.column
            if "dni" in v:                             cols["dni"]    = c
            elif "miembro" in v or "estado" in v:      cols["estado"] = c
            elif "cargo" in v:                         cols["cargo"]  = c
            elif "región" in v or "region" in v:       cols["region"] = c
            elif "provincia" in v:                     cols["prov"]   = c
            elif "distrito" in v:                      cols["dist"]   = c
            elif "mesa" in v:                          cols["mesa"]   = c
            elif "local" in v:                         cols["local"]  = c
            elif "dirección" in v or "direccion" in v: cols["dir"]    = c
            elif "referencia" in v:                    cols["ref"]    = c
            elif "tipo" in v or "voto" in v:           cols["tipo"]   = c

        # Asegurar cabeceras de columnas extra
        extras = {
            cols["local"]: "Local de Votacion",
            cols["ref"]:   "Referencia",
            cols["tipo"]:  "Tipo de Voto",
        }
        for col, titulo in extras.items():
            c = ws.cell(1, col)
            if not c.value:
                c.value     = titulo
                c.fill      = HDR_FILL
                c.font      = HDR_FONT
                c.alignment = CENTER
                ws.column_dimensions[get_column_letter(col)].width = 24

        # Filas validas
        filas = []
        for row in ws.iter_rows(min_row=2):
            v = str(row[cols["dni"] - 1].value or "").strip()
            if v.isdigit() and len(v) == 8:
                filas.append(row[0].row)

        estado["total"] = len(filas)
        log(f"Total de DNIs a consultar: {len(filas)}")

        def cel(fila, col, val="", fill=None, font=CELL_FONT, align=CENTER):
            c = ws.cell(fila, col, val)
            c.font      = font
            c.alignment = align
            c.border    = BORDER
            if fill:
                c.fill = fill

        for fila_num in filas:
            dni    = str(ws.cell(fila_num, cols["dni"]).value or "").strip()
            nombre = str(ws.cell(fila_num, cols["nombre"]).value or "").strip()
            log(f"Consultando {dni} ({nombre})...")

            res = consultar_dni(dni)

            if not res["ok"]:
                cel(fila_num, cols["estado"], "ERROR", fill=ERR_FILL)
                estado["errores"] += 1
                log(f"   [ERROR] {res['error']}")

            elif res["es_miembro"]:
                cel(fila_num, cols["estado"], "SI",               fill=YES_FILL, font=YES_FONT)
                cel(fila_num, cols["cargo"],  res["cargo"])
                cel(fila_num, cols["region"], res["region"])
                cel(fila_num, cols["prov"],   res["provincia"])
                cel(fila_num, cols["dist"],   res["distrito"])
                cel(fila_num, cols["mesa"],   res["mesa"])
                cel(fila_num, cols["local"],  res["local"],       align=LEFT)
                cel(fila_num, cols["dir"],    res["direccion"],   align=LEFT)
                cel(fila_num, cols["ref"],    res["referencia"],  align=LEFT)
                cel(fila_num, cols["tipo"],   res["tipo_voto"])
                log(f"   [SI] {res['cargo']} | Mesa {res['mesa']} | {res['distrito']}")

            else:
                cel(fila_num, cols["estado"], "NO", fill=NO_FILL, font=NO_FONT)
                log(f"   [NO] No es miembro de mesa")

            estado["done"] += 1

        wb.save(ruta_salida)
        log(f"[OK] Completado: {estado['done']} consultados, {estado['errores']} errores.")

    except Exception as e:
        log(f"[ERROR] Error critico: {e}")
    finally:
        estado["running"] = False


# ─── HTML ─────────────────────────────────────────────────────────────────────

HTML = """
<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Consulta Masiva ONPE 2026</title>
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{font-family:'Segoe UI',Arial,sans-serif;background:#f0f4f8;min-height:100vh;padding:30px 16px}
  .wrap{max-width:720px;margin:0 auto}
  h1{color:#003087;font-size:1.45rem;margin-bottom:4px;font-weight:600}
  .sub{color:#888;font-size:.85rem;margin-bottom:28px}
  .card{background:#fff;border-radius:12px;padding:26px;box-shadow:0 2px 10px rgba(0,0,0,.07);margin-bottom:18px}
  .card h2{font-size:.9rem;color:#555;margin-bottom:14px;font-weight:600;text-transform:uppercase;letter-spacing:.4px}
  .drop-zone{border:2px dashed #b8cce4;border-radius:8px;padding:30px 20px;text-align:center;cursor:pointer;transition:.2s;color:#666}
  .drop-zone:hover,.drop-zone.over{border-color:#003087;background:#f4f7ff}
  .drop-zone p{font-size:.88rem;margin-top:4px}
  .drop-zone .hint{font-size:.75rem;color:#aaa;margin-top:3px}
  input[type=file]{display:none}
  .fname{margin-top:8px;font-size:.83rem;color:#003087;font-weight:600}
  .btn{width:100%;padding:12px;border:none;border-radius:7px;font-size:.92rem;font-weight:600;cursor:pointer;transition:opacity .2s;margin-top:10px}
  .btn-primary{background:linear-gradient(135deg,#003087,#c8102e);color:#fff}
  .btn-primary:hover{opacity:.88}
  .btn-primary:disabled{opacity:.4;cursor:not-allowed}
  .btn-dl{background:#1a6f3c;color:#fff;display:none}
  .btn-dl:hover{opacity:.88}
  .prog-wrap{margin-top:14px;display:none}
  .prog-bar{height:6px;background:#dce6f0;border-radius:3px;overflow:hidden}
  .prog-fill{height:100%;background:linear-gradient(90deg,#003087,#c8102e);border-radius:3px;transition:width .35s}
  .prog-lbl{font-size:.77rem;color:#777;margin-top:4px;text-align:right}
  .log-box{background:#161b22;color:#c9d1d9;font-family:'Consolas',monospace;font-size:.76rem;
           padding:12px 14px;border-radius:7px;max-height:240px;overflow-y:auto;
           margin-top:10px;display:none;line-height:1.6}
  .log-box .ok{color:#3fb950}
  .log-box .er{color:#f85149}
  .log-box .no{color:#d29922}
  .log-box .inf{color:#79c0ff}
  .step{display:flex;gap:12px;margin-bottom:10px;align-items:flex-start}
  .snum{background:#003087;color:#fff;border-radius:50%;width:22px;height:22px;min-width:22px;
        display:flex;align-items:center;justify-content:center;font-size:.73rem;font-weight:700;margin-top:2px}
  .step p{font-size:.86rem;color:#444;line-height:1.5}
  .notice{background:#f0f4ff;border-left:3px solid #003087;border-radius:0 6px 6px 0;
          padding:10px 14px;font-size:.81rem;color:#444;margin-top:12px;line-height:1.6}
</style>
</head>
<body>
<div class="wrap">
  <h1>Consulta Masiva de Miembros de Mesa</h1>
  <p class="sub">ONPE — Elecciones Generales 2026</p>

  <div class="card">
    <h2>Como funciona</h2>
    <div class="step"><div class="snum">1</div>
      <p>Descarga la plantilla y completa los nombres y DNIs a consultar.</p></div>
    <div class="step"><div class="snum">2</div>
      <p>Sube el archivo Excel. El sistema consultara la API de la ONPE por cada DNI.</p></div>
    <div class="step"><div class="snum">3</div>
      <p>Descarga el Excel con los resultados: SI / NO, cargo, mesa, local y direccion.</p></div>
    <a href="/plantilla" class="btn btn-primary"
       style="display:inline-block;text-align:center;text-decoration:none">
      Descargar plantilla Excel
    </a>
    <div class="notice">
      Consulta directamente la API oficial de la ONPE.
      Cada DNI tarda menos de 2 segundos.
    </div>
  </div>

  <div class="card">
    <h2>Subir Excel</h2>
    <div class="drop-zone" id="dz" onclick="document.getElementById('fi').click()">
      <p>Arrastra el archivo aqui o haz clic para seleccionar</p>
      <p class="hint">.xlsx — hasta 500 filas</p>
      <div class="fname" id="fn"></div>
    </div>
    <input type="file" id="fi" accept=".xlsx" onchange="setFile(this)">

    <button class="btn btn-primary" id="btnP" onclick="iniciar()" disabled>
      Iniciar consulta
    </button>

    <div class="prog-wrap" id="pw">
      <div class="prog-bar"><div class="prog-fill" id="pf" style="width:0%"></div></div>
      <div class="prog-lbl" id="pl">0 / 0</div>
    </div>

    <div class="log-box" id="lg"></div>

    <button class="btn btn-dl" id="btnD" onclick="descargar()">
      Descargar Excel con resultados
    </button>
  </div>
</div>

<script>
let archivo = null, jobId = null, poll = null;

const dz = document.getElementById('dz');
dz.addEventListener('dragover', e => { e.preventDefault(); dz.classList.add('over'); });
dz.addEventListener('dragleave', () => dz.classList.remove('over'));
dz.addEventListener('drop', e => {
  e.preventDefault(); dz.classList.remove('over');
  const f = e.dataTransfer.files[0];
  if (f && f.name.endsWith('.xlsx')) aplicar(f);
});

function setFile(i) { if (i.files[0]) aplicar(i.files[0]); }
function aplicar(f) {
  archivo = f;
  document.getElementById('fn').textContent = f.name;
  document.getElementById('btnP').disabled = false;
  document.getElementById('btnD').style.display = 'none';
  document.getElementById('lg').style.display = 'none';
  document.getElementById('pw').style.display = 'none';
}

async function iniciar() {
  if (!archivo) return;
  const fd = new FormData();
  fd.append('file', archivo);
  document.getElementById('btnP').disabled = true;
  document.getElementById('btnP').textContent = 'Procesando...';
  document.getElementById('pw').style.display = 'block';
  document.getElementById('lg').style.display = 'block';
  document.getElementById('lg').innerHTML = '';
  document.getElementById('btnD').style.display = 'none';
  document.getElementById('pf').style.width = '0%';

  const r = await fetch('/procesar', { method: 'POST', body: fd });
  const d = await r.json();
  if (!d.ok) { alert('Error: ' + d.error); resetBtn(); return; }
  jobId = d.id;
  poll = setInterval(pollEstado, 1500);
}

async function pollEstado() {
  const d = await (await fetch('/estado')).json();
  const pct = d.total ? Math.round(d.done / d.total * 100) : 0;
  document.getElementById('pf').style.width = pct + '%';
  document.getElementById('pl').textContent = d.done + ' / ' + d.total + ' DNIs';

  const lg = document.getElementById('lg');
  lg.innerHTML = d.log.map(l => {
    let cls = 'inf';
    if (l.includes('[SI]') || l.includes('[OK]')) cls = 'ok';
    else if (l.includes('[ERROR]'))               cls = 'er';
    else if (l.includes('[NO]'))                  cls = 'no';
    return `<div class="${cls}">${l}</div>`;
  }).join('');
  lg.scrollTop = lg.scrollHeight;

  if (!d.running) {
    clearInterval(poll);
    resetBtn();
    if (d.total > 0) document.getElementById('btnD').style.display = 'block';
  }
}

function resetBtn() {
  document.getElementById('btnP').disabled = false;
  document.getElementById('btnP').textContent = 'Iniciar consulta';
}
function descargar() { window.location.href = '/descargar/' + jobId; }
</script>
</body>
</html>
"""

# ─── Directorios ──────────────────────────────────────────────────────────────

UPLOAD_DIR = "/tmp/onpe_uploads"
OUTPUT_DIR = "/tmp/onpe_outputs"
for d in [UPLOAD_DIR, OUTPUT_DIR]:
    os.makedirs(d, exist_ok=True)

PLANTILLA = os.path.join(os.path.dirname(__file__), "consulta_onpe.xlsx")


# ─── Rutas Flask ──────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/plantilla")
def plantilla():
    return send_file(PLANTILLA, as_attachment=True,
                     download_name="plantilla_consulta_onpe.xlsx")


@app.route("/procesar", methods=["POST"])
def procesar():
    global estado
    if estado["running"]:
        return jsonify({"ok": False, "error": "Ya hay un proceso en curso."})

    f = flask_request.files.get("file")
    if not f or not f.filename.endswith(".xlsx"):
        return jsonify({"ok": False, "error": "Sube un archivo .xlsx valido."})

    job_id   = str(uuid.uuid4())[:8]
    ruta_in  = os.path.join(UPLOAD_DIR, f"{job_id}_in.xlsx")
    ruta_out = os.path.join(OUTPUT_DIR, f"{job_id}_out.xlsx")
    f.save(ruta_in)

    estado["log"] = ["Iniciando..."]
    threading.Thread(target=procesar_excel,
                     args=(ruta_in, ruta_out), daemon=True).start()
    return jsonify({"ok": True, "id": job_id})


@app.route("/estado")
def get_estado():
    return jsonify(estado)


@app.route("/descargar/<job_id>")
def descargar(job_id):
    job_id = job_id.replace("..", "").replace("/", "")
    ruta   = os.path.join(OUTPUT_DIR, f"{job_id}_out.xlsx")
    if not os.path.exists(ruta):
        return "No encontrado", 404
    return send_file(ruta, as_attachment=True,
                     download_name="resultado_onpe.xlsx")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)